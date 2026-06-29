from datetime import date, timedelta

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import sessionmaker

from app.core.constants import ACTION_CONNEXION_REUSSIE, ROLE_GERANT, ROLE_VENDEUR, TYPE_ALERTE_STOCK_FAIBLE
from app.core.exceptions import PermissionRefuseeError, ValidationError
from app.database.connection import create_app_engine
from app.database.init_db import init_database
from app.database.models import Alerte, Categorie, JournalActivite, LotProduit, Produit, Utilisateur, Vente
from app.services.alerte_service import AlerteService
from app.services.auth_service import SessionUtilisateur
from app.services.rapport_service import RapportService
from app.services.vente_service import LignePanierPayload, VentePayload, VenteService


def test_historique_gerant_voit_tout_et_vendeur_uniquement_ses_ventes(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    gerant = _creer_utilisateur(SessionLocal, ROLE_GERANT, "gerant@test.local", "Gerant")
    jean = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "jean@test.local", "Jean K.")
    alice = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "alice@test.local", "Alice M.")
    vente_jean = _creer_vente(SessionLocal, jean, "Paracetamol", 1000, 2)
    vente_alice = _creer_vente(SessionLocal, alice, "Vitamine C", 750, 1)
    service = RapportService(session_factory=SessionLocal)

    toutes = service.lister_ventes(gerant)
    personnelles = service.lister_ventes(jean)

    engine.dispose()

    assert {vente.vente_id for vente in toutes} == {vente_jean, vente_alice}
    assert [vente.vente_id for vente in personnelles] == [vente_jean]


def test_synthese_dashboard_vendeur_reste_strictement_personnelle(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    jean = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "jean@test.local", "Jean K.")
    alice = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "alice@test.local", "Alice M.")
    _creer_vente(SessionLocal, jean, "Paracetamol", 1000, 2)
    _creer_vente(SessionLocal, alice, "Vitamine C", 5000, 4)
    service = RapportService(session_factory=SessionLocal)

    data = service.synthese_vendeur(jean, date_reference=date.today())

    engine.dispose()

    assert data.total == 2000
    assert data.transactions == 1
    assert data.articles == 2
    assert data.panier_moyen == 2000
    assert data.meilleure_vente == 2000
    assert [produit.produit_nom for produit in data.produits] == ["Paracetamol"]
    assert data.evolution_horaire


def test_synthese_gerant_calcule_journalier_mensuel_vendeur_et_top_produits(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    gerant = _creer_utilisateur(SessionLocal, ROLE_GERANT, "gerant@test.local", "Gerant")
    jean = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "jean@test.local", "Jean K.")
    _creer_vente(SessionLocal, jean, "Paracetamol", 1000, 2)
    _creer_vente(SessionLocal, jean, "Vitamine C", 750, 1)
    service = RapportService(session_factory=SessionLocal)

    data = service.synthese_gerant(gerant, date_reference=date.today())

    engine.dispose()

    assert data.ventes_jour == 2
    assert data.total_jour == 2750
    assert data.ventes_mois == 2
    assert data.total_mois == 2750
    assert data.panier_moyen == 1375
    assert data.vendeurs[0].vendeur_nom == "Jean K."
    assert data.vendeurs[0].total == 2750
    assert data.produits[0].produit_nom == "Paracetamol"
    assert data.produits[0].quantite == 2


def test_vendeur_refuse_rapport_global(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    vendeur = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "vendeur@test.local", "Vendeur")
    service = RapportService(session_factory=SessionLocal)

    with pytest.raises(PermissionRefuseeError):
        service.synthese_gerant(vendeur)

    engine.dispose()


def test_rapport_periode_complete_jours_categories_et_performances(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    gerant = _creer_utilisateur(SessionLocal, ROLE_GERANT, "gerant@test.local", "Gerant")
    jean = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "jean@test.local", "Jean K.")
    vente_ancienne = _creer_vente(SessionLocal, jean, "Paracetamol", 1000, 2)
    vente_recente = _creer_vente(SessionLocal, jean, "Savon medical", 1500, 1)
    _dater_et_categoriser(SessionLocal, vente_ancienne, date.today() - timedelta(days=2), "Medicaments")
    _dater_et_categoriser(SessionLocal, vente_recente, date.today(), "Soins & Hygiene")
    debut = date.today() - timedelta(days=2)
    service = RapportService(session_factory=SessionLocal)

    rapport = service.rapport_periode(
        gerant,
        date_debut=debut,
        date_fin=date.today(),
        mode="JOURNALIER",
    )

    engine.dispose()

    assert len(rapport.evolution) == (date.today() - debut).days + 1
    assert sum(item.total for item in rapport.evolution) == 3500
    assert {item.categorie_nom for item in rapport.categories} == {"Medicaments", "Soins & Hygiene"}
    assert rapport.produits_vendus == 3
    assert rapport.vendeurs[0].produits_vendus == 3
    assert rapport.vendeurs[0].panier_moyen == 1750
    assert rapport.vendeurs[0].part_ca == 100


def test_export_excel_rapport_et_validation_periode(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    gerant = _creer_utilisateur(SessionLocal, ROLE_GERANT, "gerant@test.local", "Gerant")
    vendeur = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "vendeur@test.local", "Vendeur")
    _creer_vente(SessionLocal, vendeur, "Paracetamol", 1000, 2)
    service = RapportService(session_factory=SessionLocal)
    rapport = service.rapport_periode(
        gerant,
        date_debut=date.today(),
        date_fin=date.today(),
    )

    destination = service.exporter_excel(gerant, rapport, tmp_path / "rapport.xlsx")
    workbook = load_workbook(destination)

    with pytest.raises(PermissionRefuseeError):
        service.exporter_excel(vendeur, rapport, tmp_path / "interdit.xlsx")
    with pytest.raises(ValidationError):
        service.rapport_periode(
            gerant,
            date_debut=date.today(),
            date_fin=date.today() - timedelta(days=1),
        )

    engine.dispose()

    assert destination.exists()
    assert workbook.sheetnames == [
        "Synthese",
        "Evolution",
        "Categories",
        "Vendeurs",
        "Produits vendus",
    ]
    assert workbook["Synthese"]["B5"].value == 2000
    assert workbook["Synthese"]["B5"].number_format == '#,##0 "CDF"'


def test_alertes_liste_marque_lue_et_refuse_vendeur(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    gerant = _creer_utilisateur(SessionLocal, ROLE_GERANT, "gerant@test.local", "Gerant")
    vendeur = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "vendeur@test.local", "Vendeur")
    produit_id = _creer_produit_et_alerte(SessionLocal)
    service = AlerteService(session_factory=SessionLocal)

    alertes = service.lister_alertes(gerant)
    service.marquer_lue(gerant, alerte_id=alertes[0].id)

    with SessionLocal() as session:
        alerte = session.execute(select(Alerte).where(Alerte.produit_id == produit_id)).scalar_one()

    with pytest.raises(PermissionRefuseeError):
        service.lister_alertes(vendeur)

    engine.dispose()

    assert len(alertes) == 1
    assert alerte.est_lue == 1


def test_historique_actions_gerant_lit_journal_et_vendeur_refuse(tmp_path):
    engine, SessionLocal = _create_test_session_factory(tmp_path)
    gerant = _creer_utilisateur(SessionLocal, ROLE_GERANT, "gerant@test.local", "Gerant")
    vendeur = _creer_utilisateur(SessionLocal, ROLE_VENDEUR, "vendeur@test.local", "Jean K.")
    with SessionLocal() as session:
        session.add(
            JournalActivite(
                utilisateur_id=vendeur.utilisateur_id,
                action=ACTION_CONNEXION_REUSSIE,
                table_cible="utilisateurs",
                element_id=vendeur.utilisateur_id,
                details="Connexion depuis le poste caisse.",
            )
        )
        session.commit()
    service = RapportService(session_factory=SessionLocal)

    actions = service.lister_actions(gerant, terme="connexion")
    with pytest.raises(PermissionRefuseeError):
        service.lister_actions(vendeur)

    engine.dispose()

    assert len(actions) == 1
    assert actions[0].utilisateur_nom == "Jean K."
    assert actions[0].module == "Vendeurs"


def _create_test_session_factory(tmp_path):
    database_path = tmp_path / "salmospharm.sqlite3"
    engine = create_app_engine(database_path)
    init_database(database_engine=engine)
    return engine, sessionmaker(bind=engine, autoflush=False, autocommit=False, expire_on_commit=False, future=True)


def _creer_utilisateur(SessionLocal, role: str, email: str, nom: str) -> SessionUtilisateur:
    with SessionLocal() as session:
        utilisateur = Utilisateur(nom=nom, email=email, mot_de_passe_hash="hash", role=role, actif=1)
        session.add(utilisateur)
        session.commit()
        return SessionUtilisateur(utilisateur_id=utilisateur.id, nom=utilisateur.nom, identifiant=email, role=role)


def _creer_vente(SessionLocal, vendeur: SessionUtilisateur, produit_nom: str, prix: int, quantite: int) -> int:
    produit_id = _creer_produit(SessionLocal, produit_nom, prix)
    _creer_lot(SessionLocal, produit_id, quantite + 3)
    result = VenteService(session_factory=SessionLocal).valider_vente(
        vendeur,
        VentePayload(lignes=[LignePanierPayload(produit_id=produit_id, quantite=quantite)], montant_recu=prix * quantite),
        date_reference=date.today(),
    )
    with SessionLocal() as session:
        vente = session.get(Vente, result.vente_id)
        vente.cree_le = f"{date.today().isoformat()} 10:00:00"
        session.commit()
    return result.vente_id


def _creer_produit(SessionLocal, nom: str, prix: int) -> int:
    with SessionLocal() as session:
        produit = Produit(nom=nom, code_barres=f"{nom}-{prix}", prix_vente=prix, stock_minimum=0, actif=1)
        session.add(produit)
        session.commit()
        return produit.id


def _creer_lot(SessionLocal, produit_id: int, quantite: int) -> int:
    with SessionLocal() as session:
        lot = LotProduit(produit_id=produit_id, numero_lot=f"LOT-{produit_id}", quantite=quantite, prix_achat=500, date_expiration="2027-01-01")
        session.add(lot)
        session.commit()
        return lot.id


def _creer_produit_et_alerte(SessionLocal) -> int:
    produit_id = _creer_produit(SessionLocal, "Alerte produit", 500)
    with SessionLocal() as session:
        alerte = Alerte(
            produit_id=produit_id,
            lot_id=None,
            type_alerte=TYPE_ALERTE_STOCK_FAIBLE,
            message="Stock faible pour Alerte produit.",
            est_lue=0,
        )
        session.add(alerte)
        session.commit()
        return produit_id


def _dater_et_categoriser(SessionLocal, vente_id: int, jour: date, categorie_nom: str) -> None:
    with SessionLocal() as session:
        vente = session.get(Vente, vente_id)
        vente.cree_le = f"{jour.isoformat()} 10:00:00"
        produit = vente.lignes[0].produit
        categorie = Categorie(nom=categorie_nom)
        session.add(categorie)
        session.flush()
        produit.categorie_id = categorie.id
        session.commit()
