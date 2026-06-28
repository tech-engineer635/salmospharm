from datetime import date, timedelta

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import sessionmaker

from app.core.constants import ACTION_EXPORT_EXCEL, ROLE_GERANT, ROLE_VENDEUR
from app.core.exceptions import PermissionRefuseeError, ValidationError
from app.database.connection import create_app_engine
from app.database.init_db import init_database
from app.database.models import (
    Categorie,
    JournalActivite,
    LigneVente,
    LotProduit,
    Produit,
    Utilisateur,
    Vente,
)
from app.services.auth_service import SessionUtilisateur
from app.services.produit_service import ProduitService
from app.services.rapport_service import RapportService
from app.services.stock_service import StockService


def test_exports_produits_et_stock_creent_des_classeurs_types_et_surs(tmp_path):
    engine, SessionLocal = _environment(tmp_path)
    gerant = _user(SessionLocal, ROLE_GERANT, "gerant-excel@test.local")
    categorie_id = _catalogue(SessionLocal)
    produit_service = ProduitService(session_factory=SessionLocal)
    stock_service = StockService(session_factory=SessionLocal)

    products_path = produit_service.exporter_excel(
        gerant,
        destination=tmp_path / "produits.xlsx",
        categorie_id=categorie_id,
        statut="ACTIFS",
    )
    stock_path = stock_service.exporter_excel(gerant, tmp_path / "stock.xlsx")
    products = load_workbook(products_path)
    stock = load_workbook(stock_path)

    with SessionLocal() as session:
        export_actions = session.scalar(
            select(func.count())
            .select_from(JournalActivite)
            .where(JournalActivite.action == ACTION_EXPORT_EXCEL)
        )

    engine.dispose()

    assert products["Produits"]["B2"].value == "'=Produit test"
    assert products["Produits"]["E2"].value == 1500
    assert products["Produits"]["E2"].number_format == '#,##0 "CDF"'
    assert stock["Stock"]["F2"].value == 900
    assert stock["Stock"]["H2"].value == "Expire"
    assert export_actions == 2


def test_export_ventes_respecte_filtres_et_ne_depend_pas_limite_affichage(tmp_path):
    engine, SessionLocal = _environment(tmp_path)
    gerant = _user(SessionLocal, ROLE_GERANT, "gerant-ventes@test.local")
    vendeur = _user(SessionLocal, ROLE_VENDEUR, "vendeur-ventes@test.local")
    _sales(SessionLocal, vendeur.utilisateur_id, count=105)
    service = RapportService(session_factory=SessionLocal)

    path = service.exporter_ventes_excel(
        gerant,
        destination=tmp_path / "ventes.xlsx",
        terme="VTE-",
        date_debut=date.today(),
        date_fin=date.today(),
    )
    workbook = load_workbook(path)
    sheet = workbook["Ventes"]

    engine.dispose()

    assert sheet.max_row == 106
    assert sheet["E2"].data_type == "n"
    assert sheet["E2"].number_format == '#,##0 "CDF"'
    assert sheet["H2"].value == "VALIDEE"


def test_tous_les_exports_refusent_vendeur_et_export_vide_est_clair(tmp_path):
    engine, SessionLocal = _environment(tmp_path)
    vendeur = _user(SessionLocal, ROLE_VENDEUR, "vendeur-refuse@test.local")
    product_service = ProduitService(session_factory=SessionLocal)
    stock_service = StockService(session_factory=SessionLocal)
    report_service = RapportService(session_factory=SessionLocal)

    with pytest.raises(PermissionRefuseeError):
        product_service.exporter_excel(vendeur, destination=tmp_path / "p.xlsx")
    with pytest.raises(PermissionRefuseeError):
        stock_service.exporter_excel(vendeur, tmp_path / "s.xlsx")
    with pytest.raises(PermissionRefuseeError):
        report_service.exporter_ventes_excel(vendeur, destination=tmp_path / "v.xlsx")

    gerant = _user(SessionLocal, ROLE_GERANT, "gerant-vide@test.local")
    with pytest.raises(ValidationError, match="Aucune donnee"):
        report_service.exporter_ventes_excel(
            gerant,
            destination=tmp_path / "vide.xlsx",
            terme="INEXISTANT",
        )

    engine.dispose()


def _environment(tmp_path):
    database_path = tmp_path / "salmospharm.sqlite3"
    engine = create_app_engine(database_path)
    init_database(database_engine=engine)
    SessionLocal = sessionmaker(
        bind=engine,
        autoflush=False,
        autocommit=False,
        expire_on_commit=False,
        future=True,
    )
    return engine, SessionLocal


def _user(SessionLocal, role: str, email: str) -> SessionUtilisateur:
    with SessionLocal() as session:
        user = Utilisateur(
            nom=email.split("@")[0],
            email=email,
            mot_de_passe_hash="hash",
            role=role,
            actif=1,
        )
        session.add(user)
        session.commit()
        return SessionUtilisateur(user.id, user.nom, user.email, user.role)


def _catalogue(SessionLocal) -> int:
    with SessionLocal() as session:
        category = Categorie(nom="Export Excel")
        session.add(category)
        session.flush()
        product = Produit(
            categorie_id=category.id,
            nom="=Produit test",
            code_barres="EXCEL-001",
            prix_vente=1500,
            stock_minimum=5,
            actif=1,
        )
        session.add(product)
        session.flush()
        session.add(
            LotProduit(
                produit_id=product.id,
                numero_lot="LOT-EXPIRE",
                quantite=3,
                prix_achat=900,
                date_expiration=(date.today() - timedelta(days=1)).isoformat(),
            )
        )
        session.commit()
        return category.id


def _sales(SessionLocal, seller_id: int, *, count: int) -> None:
    with SessionLocal() as session:
        product = Produit(
            nom="Produit ventes",
            code_barres="VENTES-001",
            prix_vente=1000,
            stock_minimum=0,
            actif=1,
        )
        session.add(product)
        session.flush()
        for index in range(count):
            sale = Vente(
                numero_vente=f"VTE-{index + 1:06d}",
                vendeur_id=seller_id,
                total=1000,
                montant_recu=1500,
                statut="VALIDEE",
            )
            session.add(sale)
            session.flush()
            session.add(
                LigneVente(
                    vente_id=sale.id,
                    produit_id=product.id,
                    quantite=1,
                    prix_unitaire=1000,
                    sous_total=1000,
                )
            )
        session.commit()
