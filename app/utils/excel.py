"""Outils communs pour produire des classeurs Excel lisibles et surs."""

from __future__ import annotations

import os
import uuid
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.core.exceptions import ValidationError


CDF_FORMAT = '#,##0 "CDF"'
DATE_FORMAT = "dd/mm/yyyy"
DATETIME_FORMAT = "dd/mm/yyyy hh:mm"
PERCENT_FORMAT = "0.0%"
HEADER_FILL = "0B3567"


def creer_classeur_tableau(
    *,
    titre_feuille: str,
    entetes: Sequence[str],
    lignes: Iterable[Sequence[object]],
    colonnes_cdf: Sequence[int] = (),
    colonnes_date: Sequence[int] = (),
    colonnes_datetime: Sequence[int] = (),
    colonnes_pourcentage: Sequence[int] = (),
) -> tuple[Workbook, int]:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = titre_feuille[:31]
    sheet.append(list(entetes))
    row_count = 0
    for values in lignes:
        sheet.append([_securiser_cellule(value) for value in values])
        row_count += 1
    if row_count == 0:
        raise ValidationError("Aucune donnee ne correspond aux filtres selectionnes.")

    styliser_feuille(
        sheet,
        colonnes_cdf=colonnes_cdf,
        colonnes_date=colonnes_date,
        colonnes_datetime=colonnes_datetime,
        colonnes_pourcentage=colonnes_pourcentage,
    )
    table = Table(displayName=f"Tableau{uuid.uuid4().hex[:10]}", ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    return workbook, row_count


def styliser_feuille(
    sheet,
    *,
    colonnes_cdf: Sequence[int] = (),
    colonnes_date: Sequence[int] = (),
    colonnes_datetime: Sequence[int] = (),
    colonnes_pourcentage: Sequence[int] = (),
) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 24

    formats = {
        **{index: CDF_FORMAT for index in colonnes_cdf},
        **{index: DATE_FORMAT for index in colonnes_date},
        **{index: DATETIME_FORMAT for index in colonnes_datetime},
        **{index: PERCENT_FORMAT for index in colonnes_pourcentage},
    }
    for index, number_format in formats.items():
        for cell in sheet.iter_cols(min_col=index, max_col=index, min_row=2):
            for item in cell:
                item.number_format = number_format

    for column in sheet.columns:
        width = min(44, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width


def enregistrer_classeur(workbook: Workbook, destination: str | Path) -> Path:
    path = Path(destination)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    try:
        workbook.save(temporary)
        os.replace(temporary, path)
    except (OSError, PermissionError) as exc:
        temporary.unlink(missing_ok=True)
        raise ValidationError(
            "Impossible d'enregistrer le fichier Excel. Fermez-le s'il est deja ouvert."
        ) from exc
    return path


def convertir_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def convertir_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value[:10])
    except ValueError:
        return None


def _securiser_cellule(value: object) -> object:
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value
