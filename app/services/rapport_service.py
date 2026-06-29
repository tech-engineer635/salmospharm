"""Rapports et historiques calcules depuis les ventes validees."""

from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

from app.core.constants import ACTION_EXPORT_EXCEL, ROLE_GERANT
from app.core.exceptions import ValidationError
from app.core.paths import get_exports_dir
from app.core.permissions import (
    PERMISSION_CONSULTER_HISTORIQUE_PERSONNEL,
    PERMISSION_CONSULTER_HISTORIQUE_SYSTEME,
    PERMISSION_CONSULTER_RAPPORTS_GLOBAUX,
    PERMISSION_CONSULTER_TOUTES_VENTES,
    PERMISSION_EXPORTER_DONNEES,
    exiger_permission,
)
from app.database.connection import create_session
from app.repositories.rapport_repository import RapportRepository
from app.services.auth_service import SessionUtilisateur
from app.services.journal_service import JournalService
from app.utils.excel import (
    convertir_datetime,
    creer_classeur_tableau,
    enregistrer_classeur,
    styliser_feuille,
)


@dataclass(frozen=True)
class VenteHistoriqueItem:
    vente_id: int
    numero_vente: str
    vendeur_id: int | None
    vendeur_nom: str
    date_vente: str
    total: int
    montant_recu: int
    monnaie_rendue: int
    articles: int


@dataclass(frozen=True)
class RapportEvolutionItem:
    date_vente: date
    transactions: int
    total: int


@dataclass(frozen=True)
class RapportCategorieItem:
    categorie_nom: str
    quantite: int
    total: int


@dataclass(frozen=True)
class RapportVendeurItem:
    vendeur_id: int | None
    vendeur_nom: str
    ventes: int
    total: int
    produits_vendus: int = 0
    panier_moyen: int = 0
    part_ca: float = 0.0
    evolution: float = 0.0


@dataclass(frozen=True)
class ProduitVenduItem:
    produit_id: int
    produit_nom: str
    quantite: int
    total: int


@dataclass(frozen=True)
class RapportSynthese:
    ventes_jour: int
    total_jour: int
    ventes_mois: int
    total_mois: int
    panier_moyen: int
    vendeurs: list[RapportVendeurItem]
    produits: list[ProduitVenduItem]
    date_debut: date | None = None
    date_fin: date | None = None
    produits_vendus: int = 0
    evolution: list[RapportEvolutionItem] | None = None
    categories: list[RapportCategorieItem] | None = None
    tendance_ca: float = 0.0
    tendance_transactions: float = 0.0
    tendance_panier: float = 0.0
    tendance_produits: float = 0.0
    mode: str = "JOURNALIER"


@dataclass(frozen=True)
class JournalActionItem:
    journal_id: int
    date_action: str
    utilisateur_nom: str
    action: str
    details: str
    module: str


class RapportService:
    """Valide les permissions et transforme les aggregations en rapports UI."""

    def __init__(
        self,
        session_factory=create_session,
        rapport_repository: RapportRepository | None = None,
        journal_service: JournalService | None = None,
    ) -> None:
        self._session_factory = session_factory
        self._rapport_repository = rapport_repository or RapportRepository()
        self._journal_service = journal_service or JournalService()

    def lister_ventes(
        self,
        utilisateur: SessionUtilisateur,
        *,
        terme: str = "",
        date_debut: date | None = None,
        date_fin: date | None = None,
        limit: int = 100,
    ) -> list[VenteHistoriqueItem]:
        """Retourne l'historique autorise : tout pour gerant, personnel pour vendeur."""
        if utilisateur.role == ROLE_GERANT:
            exiger_permission(utilisateur.role, PERMISSION_CONSULTER_TOUTES_VENTES)
            vendeur_id = None
        else:
            exiger_permission(utilisateur.role, PERMISSION_CONSULTER_HISTORIQUE_PERSONNEL)
            vendeur_id = utilisateur.utilisateur_id

        with self._session_factory() as session:
            rows = self._rapport_repository.lister_ventes(
                session,
                vendeur_id=vendeur_id,
                terme=terme,
                date_debut=date_debut,
                date_fin=date_fin,
                limit=limit,
            )
        return [
            VenteHistoriqueItem(
                vente_id=row[0],
                numero_vente=row[1],
                vendeur_id=row[2],
                vendeur_nom=row[3],
                date_vente=_format_datetime(row[4]),
                total=row[5],
                montant_recu=row[6],
                monnaie_rendue=row[6] - row[5],
                articles=int(row[7]),
            )
            for row in rows
        ]

    def synthese_gerant(
        self,
        utilisateur: SessionUtilisateur,
        *,
        date_reference: date | None = None,
        date_debut: date | None = None,
        date_fin: date | None = None,
        mode: str = "JOURNALIER",
    ) -> RapportSynthese:
        """Calcule les indicateurs globaux sur une periode inclusive."""

        reference = date_reference or date.today()
        debut = date_debut or reference
        fin = date_fin or reference
        data = self.rapport_periode(
            utilisateur, date_debut=debut, date_fin=fin, mode=mode
        )
        if date_debut is None and date_fin is None:
            with self._session_factory() as session:
                ventes_mois, total_mois, _ = self._rapport_repository.resume_periode(
                    session, reference.replace(day=1), reference
                )
            data = replace(data, ventes_mois=ventes_mois, total_mois=total_mois)
        return data

    def rapport_periode(
        self,
        utilisateur: SessionUtilisateur,
        *,
        date_debut: date,
        date_fin: date,
        mode: str = "JOURNALIER",
    ) -> RapportSynthese:
        """Construit le rapport et sa comparaison a la periode precedente."""

        exiger_permission(utilisateur.role, PERMISSION_CONSULTER_RAPPORTS_GLOBAUX)
        _valider_periode(date_debut, date_fin)
        mode_normalise = mode.strip().upper()
        if mode_normalise not in {"JOURNALIER", "MENSUEL", "VENDEUR"}:
            raise ValidationError("Type de rapport inconnu.")

        duree = (date_fin - date_debut).days + 1
        precedent_fin = date_debut - timedelta(days=1)
        precedent_debut = precedent_fin - timedelta(days=duree - 1)

        with self._session_factory() as session:
            ventes, total, produits_vendus = self._rapport_repository.resume_periode(
                session, date_debut, date_fin
            )
            ventes_prec, total_prec, produits_prec = self._rapport_repository.resume_periode(
                session, precedent_debut, precedent_fin
            )
            evolution_rows = self._rapport_repository.evolution_journaliere(
                session, date_debut, date_fin
            )
            categories_rows = self._rapport_repository.repartition_categories(
                session, date_debut, date_fin
            )
            vendeur_rows = self._rapport_repository.performances_vendeurs(
                session, date_debut, date_fin
            )
            vendeurs_prec = self._rapport_repository.totaux_vendeurs(
                session, precedent_debut, precedent_fin
            )
            produits_rows = self._rapport_repository.produits_plus_vendus(
                session, date_debut, date_fin
            )

        panier = int(total / ventes) if ventes else 0
        panier_prec = int(total_prec / ventes_prec) if ventes_prec else 0
        evolution_map = {
            date.fromisoformat(jour): (transactions, montant)
            for jour, transactions, montant in evolution_rows
        }
        evolution = []
        jour = date_debut
        while jour <= date_fin:
            transactions, montant = evolution_map.get(jour, (0, 0))
            evolution.append(RapportEvolutionItem(jour, transactions, montant))
            jour += timedelta(days=1)

        vendeurs = [
            RapportVendeurItem(
                vendeur_id=vendeur_id,
                vendeur_nom=nom,
                ventes=transactions,
                total=montant,
                produits_vendus=quantite,
                panier_moyen=int(montant / transactions) if transactions else 0,
                part_ca=(montant * 100 / total) if total else 0.0,
                evolution=_variation(montant, vendeurs_prec.get(vendeur_id, 0)),
            )
            for vendeur_id, nom, transactions, montant, quantite in vendeur_rows
        ]

        return RapportSynthese(
            ventes_jour=ventes,
            total_jour=total,
            ventes_mois=ventes,
            total_mois=total,
            panier_moyen=panier,
            vendeurs=vendeurs,
            produits=[
                ProduitVenduItem(int(row[0]), str(row[1]), int(row[2]), int(row[3]))
                for row in produits_rows
            ],
            date_debut=date_debut,
            date_fin=date_fin,
            produits_vendus=produits_vendus,
            evolution=evolution,
            categories=[
                RapportCategorieItem(str(row[0]), int(row[1]), int(row[2]))
                for row in categories_rows
            ],
            tendance_ca=_variation(total, total_prec),
            tendance_transactions=_variation(ventes, ventes_prec),
            tendance_panier=_variation(panier, panier_prec),
            tendance_produits=_variation(produits_vendus, produits_prec),
            mode=mode_normalise,
        )

    def exporter_excel(
        self,
        utilisateur: SessionUtilisateur,
        rapport: RapportSynthese,
        destination: str | Path | None = None,
    ) -> Path:
        """Exporte le rapport affiche dans un classeur Excel CDF."""

        exiger_permission(utilisateur.role, PERMISSION_EXPORTER_DONNEES)
        debut = rapport.date_debut or date.today()
        fin = rapport.date_fin or debut
        path = Path(destination) if destination else (
            get_exports_dir()
            / f"rapport_{rapport.mode.lower()}_{debut.isoformat()}_{fin.isoformat()}.xlsx"
        )
        if path.suffix.lower() != ".xlsx":
            path = path.with_suffix(".xlsx")
        workbook = _creer_classeur(rapport)
        path = enregistrer_classeur(workbook, path)
        with self._session_factory() as session:
            self._journal_service.journaliser(
                session,
                action=ACTION_EXPORT_EXCEL,
                utilisateur_id=utilisateur.utilisateur_id,
                table_cible="ventes",
                details=(
                    f"Export Excel rapport {rapport.mode.lower()} : {path.name}, "
                    f"periode {debut.isoformat()} au {fin.isoformat()}."
                ),
            )
            session.commit()
        return path

    def exporter_ventes_excel(
        self,
        utilisateur: SessionUtilisateur,
        *,
        destination: str | Path | None = None,
        terme: str = "",
        date_debut: date | None = None,
        date_fin: date | None = None,
    ) -> Path:
        exiger_permission(utilisateur.role, PERMISSION_EXPORTER_DONNEES)
        if date_debut is not None and date_fin is not None:
            _valider_periode(date_debut, date_fin)
        with self._session_factory() as session:
            rows = self._rapport_repository.lister_ventes(
                session,
                vendeur_id=None,
                terme=terme,
                date_debut=date_debut,
                date_fin=date_fin,
                limit=None,
            )
            workbook, count = creer_classeur_tableau(
                titre_feuille="Ventes",
                entetes=(
                    "Numero de vente",
                    "Date et heure",
                    "Vendeur",
                    "Nombre d'articles",
                    "Total (CDF)",
                    "Montant recu (CDF)",
                    "Monnaie rendue (CDF)",
                    "Statut",
                ),
                lignes=(
                    (
                        row[1],
                        convertir_datetime(row[4]),
                        row[3],
                        int(row[7]),
                        int(row[5]),
                        int(row[6]),
                        int(row[6]) - int(row[5]),
                        "VALIDEE",
                    )
                    for row in rows
                ),
                colonnes_cdf=(5, 6, 7),
                colonnes_datetime=(2,),
            )
            start_label = date_debut.isoformat() if date_debut else "debut"
            end_label = date_fin.isoformat() if date_fin else date.today().isoformat()
            path = enregistrer_classeur(
                workbook,
                destination
                or get_exports_dir() / f"ventes_{start_label}_{end_label}.xlsx",
            )
            self._journal_service.journaliser(
                session,
                action=ACTION_EXPORT_EXCEL,
                utilisateur_id=utilisateur.utilisateur_id,
                table_cible="ventes",
                details=f"Export Excel ventes : {path.name}, {count} ligne(s).",
            )
            session.commit()
        return path

    def lister_actions(
        self,
        utilisateur: SessionUtilisateur,
        *,
        terme: str = "",
        limit: int = 100,
        date_action: date | None = None,
        utilisateur_nom: str = "",
        action: str = "",
    ) -> list[JournalActionItem]:
        """Retourne le journal systeme reserve au gerant."""
        exiger_permission(utilisateur.role, PERMISSION_CONSULTER_HISTORIQUE_SYSTEME)
        with self._session_factory() as session:
            rows = self._rapport_repository.lister_actions(
                session,
                terme=terme,
                limit=limit,
                date_action=date_action,
                utilisateur_nom=utilisateur_nom,
                action=action,
            )
        return [
            JournalActionItem(
                journal_id=row[0],
                date_action=_format_datetime(row[1]),
                utilisateur_nom=row[2],
                action=row[3],
                details=row[4],
                module=_module_label(row[5]),
            )
            for row in rows
        ]


def _valider_periode(date_debut: date, date_fin: date) -> None:
    if date_debut > date_fin:
        raise ValidationError("La date de debut doit preceder la date de fin.")
    if (date_fin - date_debut).days > 366:
        raise ValidationError("La periode du rapport ne peut pas depasser 12 mois.")


def _variation(valeur: int, precedente: int) -> float:
    if precedente == 0:
        return 100.0 if valeur > 0 else 0.0
    return (valeur - precedente) * 100 / precedente


def _creer_classeur(rapport: RapportSynthese) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Synthese"
    summary.append(["SALMOSPHARM 133", "Rapport", rapport.mode.title()])
    summary.append(["Periode", str(rapport.date_debut or ""), str(rapport.date_fin or "")])
    summary.append([])
    summary.append(["Indicateur", "Valeur", "Unite"])
    summary.append(["Chiffre d'affaires", rapport.total_jour, "CDF"])
    summary.append(["Transactions", rapport.ventes_jour, ""])
    summary.append(["Panier moyen", rapport.panier_moyen, "CDF"])
    summary.append(["Produits vendus", rapport.produits_vendus, ""])

    evolution = workbook.create_sheet("Evolution")
    evolution.append(["Date", "Transactions", "Ventes (CDF)"])
    for item in rapport.evolution or []:
        evolution.append([item.date_vente.isoformat(), item.transactions, item.total])

    categories = workbook.create_sheet("Categories")
    categories.append(["Categorie", "Produits vendus", "Ventes (CDF)", "Part du CA"])
    for item in rapport.categories or []:
        part = item.total / rapport.total_jour if rapport.total_jour else 0
        categories.append([item.categorie_nom, item.quantite, item.total, part])

    vendeurs = workbook.create_sheet("Vendeurs")
    vendeurs.append(
        ["Vendeur", "Transactions", "Ventes (CDF)", "Panier moyen (CDF)", "Produits vendus", "Part du CA", "Evolution"]
    )
    for item in rapport.vendeurs:
        vendeurs.append(
            [
                item.vendeur_nom,
                item.ventes,
                item.total,
                item.panier_moyen,
                item.produits_vendus,
                item.part_ca / 100,
                item.evolution / 100,
            ]
        )

    produits = workbook.create_sheet("Produits vendus")
    produits.append(["Produit", "Quantite vendue", "Ventes (CDF)"])
    for item in rapport.produits:
        produits.append([item.produit_nom, item.quantite, item.total])

    for sheet in workbook.worksheets:
        styliser_feuille(sheet)
    for cell in summary["B"][4:7]:
        if cell.row in {5, 7}:
            cell.number_format = '#,##0 "CDF"'
    for cell in evolution["C"][1:]:
        cell.number_format = '#,##0 "CDF"'
    for cell in categories["C"][1:]:
        cell.number_format = '#,##0 "CDF"'
    for column in ("C", "D"):
        for cell in vendeurs[column][1:]:
            cell.number_format = '#,##0 "CDF"'
    for cell in produits["C"][1:]:
        cell.number_format = '#,##0 "CDF"'
    for cell in categories["D"][1:]:
        cell.number_format = "0.0%"
    for cell in vendeurs["F"][1:] + vendeurs["G"][1:]:
        cell.number_format = "0.0%"
    return workbook


def _format_datetime(value: str) -> str:
    return value[:16].replace("-", "/") if value else ""


def _module_label(value: str) -> str:
    labels = {
        "ventes": "Ventes",
        "produits": "Produits",
        "lots_produits": "Stock",
        "mouvements_stock": "Stock",
        "utilisateurs": "Vendeurs",
        "parametres": "Parametres",
    }
    return labels.get(value or "", "Systeme")
